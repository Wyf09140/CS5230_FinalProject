import pandas as pd
import json
import time
import os
from openpyxl import load_workbook
from groq import Groq

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL = 'Download_data/Career Drive Project Data Sources.xlsx'
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
MODEL  = "llama-3.3-70b-versatile"

# 21个职业的 SOC codes（来自 D1）
ALL_SOC = {
    "11-3013": "Facilities Managers",
    "11-9021": "Construction Managers",
    "11-9041": "Architectural and Engineering Managers",
    "13-1082": "Project Management Specialists",
    "17-2051": "Civil Engineers",
    "17-2051.01": "Transportation Engineers",
    "17-3022": "Civil Engineering Technologists and Technicians",
    "19-4044": "Hydrologists",
    "47-1011": "First-Line Supervisors of Construction Trades",
    "47-2011": "Boilermakers",
    "47-2031": "Carpenters",
    "47-2051": "Cement Masons and Concrete Finishers",
    "47-2061": "Construction Laborers",
    "47-2071": "Paving, Surfacing, and Tamping Equipment Operators",
    "47-2073": "Operating Engineers and Other Construction Equipment Operators",
    "47-2111": "Electricians",
    "47-2152": "Plumbers, Pipefitters, and Steamfitters",
    "47-2171": "Reinforcing Iron and Rebar Workers",
    "47-2181": "Roofers",
    "47-2221": "Structural Iron and Steel Workers",
    "53-7021": "Crane and Tower Operators",
}

# General Contractors → 全量蓝领 SOC
GC_SOC = [k for k in ALL_SOC if k.startswith("47") or k in ("11-9021","11-1011","53-7021")]

# ── 1. 读取 AGC Member List ───────────────────────────────────────────────────
wb = load_workbook(EXCEL)
ws_agc = wb['AGC Member list ']

rows = []
for row in ws_agc.iter_rows(min_row=4, values_only=True):
    t, name = row[0], row[1]
    if t in ('General Contractors', 'Specialty Contractors') and name:
        rows.append({'agc_type': t.strip(), 'employer_name': name.strip()})

df_agc = pd.DataFrame(rows)
print(f"AGC: {len(df_agc)} rows — GC:{(df_agc.agc_type=='General Contractors').sum()}, SC:{(df_agc.agc_type=='Specialty Contractors').sum()}")

# ── 2. 读取 DOT Prequal URLs ──────────────────────────────────────────────────
ws_dot = wb['JobBoards -DOT Prequal List']
dot_urls = {}
for row in ws_dot.iter_rows(min_row=2):
    company = row[0].value
    cell_b  = row[1]
    url = cell_b.hyperlink.target if cell_b.hyperlink else None
    # 排除纯文字备注（无超链接）
    if company and url:
        dot_urls[company.strip().lower()] = url

print(f"DOT URLs: {len(dot_urls)} entries")

# ── 3. LLM SOC Mapping for Specialty Contractors ─────────────────────────────
soc_list_str = "\n".join([f"{k}: {v}" for k, v in ALL_SOC.items()])

def get_soc_mapping(company_name: str) -> list:
    prompt = f"""You are helping map Maine construction companies to occupation SOC codes.

Company: "{company_name}"

Available SOC codes:
{soc_list_str}

Task:
1. Based on the company name, infer what construction trades or specialties this company likely performs.
2. Return ONLY the SOC codes that are a strong match (1-5 codes max).
3. If the company name gives no clear signal, return ["GENERAL"].

Respond with ONLY a JSON array of SOC code strings. No explanation.
Examples:
- "Mancini Electric" → ["47-2111"]
- "Black Bear Crane" → ["53-7021"]
- "CMI Concrete Professionals" → ["47-2051"]
- "Gorham Sand & Gravel" → ["47-2061", "47-2071", "47-2073"]
- "Carver Carpentry & Excavation" → ["47-2031", "47-2061"]
- "Hird Contracting" → ["GENERAL"]
"""
    resp = client.chat.completions.create(
        model=MODEL,
        max_tokens=200,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.choices[0].message.content.strip()
    try:
        codes = json.loads(raw)
        # validate
        valid = [c for c in codes if c in ALL_SOC or c == "GENERAL"]
        return valid if valid else ["GENERAL"]
    except:
        return ["GENERAL"]

# 只对 Specialty Contractors 跑 LLM
sc_df = df_agc[df_agc.agc_type == 'Specialty Contractors'].copy()
print(f"\nRunning LLM mapping for {len(sc_df)} Specialty Contractors...")

CACHE_FILE = 'Output_data/soc_mapping_cache.json'

# 读缓存
if os.path.exists(CACHE_FILE):
    with open(CACHE_FILE, 'r') as f:
        soc_results = json.load(f)
    print(f"Loaded cache: {len(soc_results)} entries")
else:
    soc_results = {}

# 只跑没有缓存的公司
to_map = [row for row in sc_df.itertuples() if row.employer_name not in soc_results]
print(f"To map: {len(to_map)} companies (cached: {len(soc_results)})")

for i, row in enumerate(to_map):
    codes = get_soc_mapping(row.employer_name)
    soc_results[row.employer_name] = codes
    print(f"  {i+1}/{len(to_map)} {row.employer_name} → {codes}")
    time.sleep(0.3)

# 保存缓存
with open(CACHE_FILE, 'w') as f:
    json.dump(soc_results, f, indent=2)

print("\nLLM mapping done. Cache saved.")

# ── 4. 展开 SOC mapping → 每行一个 soc_code ───────────────────────────────────
import urllib.parse

records = []

# GC_SOC 修正：确保包含 11-9021
GC_SOC_FINAL = [k for k in ALL_SOC if k.startswith("47") or k in ("11-9021", "53-7021")]

for _, row in df_agc.iterrows():
    name = row['employer_name']
    atype = row['agc_type']

    # 确定 SOC codes
    if atype == 'General Contractors':
        soc_codes = GC_SOC_FINAL
    else:
        raw_codes = soc_results.get(name, ['GENERAL'])
        if raw_codes == ['GENERAL']:
            soc_codes = GC_SOC_FINAL  # fallback
        else:
            soc_codes = raw_codes

    # 查 DOT URL（严格匹配：公司名规范化后完全匹配）
    def normalize(s):
        return s.lower().strip().rstrip('.,inc lc').strip()

    name_lower = normalize(name)
    matched_url = None
    for dot_name, dot_url in dot_urls.items():
        if normalize(dot_name) == name_lower:
            matched_url = dot_url
            break
        # 次级匹配：dot_name 是 agc_name 的子串（至少6字符）
        dn = normalize(dot_name)
        if len(dn) >= 6 and dn in name_lower:
            matched_url = dot_url
            break

    for soc in soc_codes:
        occ_title = ALL_SOC.get(soc, '')
        # Indeed fallback URL
        indeed_url = f"https://www.indeed.com/jobs?q={urllib.parse.quote(occ_title)}&l=Maine"

        records.append({
            'soc_code':      soc,
            'occ_title':     occ_title,
            'employer_name': name,
            'agc_type':      atype,
            'url':           matched_url if matched_url else indeed_url,
            'url_type':      'company_site' if matched_url else 'indeed_fallback',
            'source_type':   'local_employer',
        })

# ── 5. Job board URLs (hardcoded — no hyperlinks in sheet) ───────────────────
jb_urls = {
    # Maine Specific
    'MaineDOT Careers':         'https://maine.wd5.myworkdayjobs.com/Executive?q=Department+of+Transportation',
    'Maine Turnpike Authority':  'https://www.maineturnpike.com/About-MTA/Employment.aspx',
    'Maine.gov Job Board':       'https://www.maine.gov/jobs/',
    'AGC Maine':                 'https://www.agcmaine.org/career-center/',
    'NSPE-Maine':                'https://www.nspe.org/resources/career-center',
    'Cianbro':                   'https://www.cianbro.com/careers-list',
    'Sargent':                   'https://www.sargentcorp.com/careers',
    'Reed & Reed':               'https://www.reedreed.com/careers',
    'Sebago Technics':           'https://www.sebagotechnics.com/careers',
    'Haley Ward':                'https://www.haleyward.com/careers',
    'Live-and-Work-in-Maine':    'https://www.liveandworkinmaine.com/jobs/',
    'Maine Hire-A-Vet':          'https://www.maine.gov/labor/hire_a_vet/',
    # National
    'ConstructionJobs.com':      'https://www.constructionjobs.com/',
    'iHireConstruction':         'https://www.ihireconstruction.com/',
    'ASCE Career Connections':   'https://careers.asce.org/',
    'EngineerJobs.com':          'https://www.engineerjobs.com/',
    'CMAA Career HQ':            'https://jobs.constructionmanagement.org/',
    'AASHTO Jobs':               'https://jobs.transportation.org/',
    'ITE Hire':                  'https://jobs.ite.org/',
    'RoadTechs':                 'https://www.roadtechs.com/wwwboard/',
    'Michael Page Construction': 'https://www.michaelpage.com/jobs/construction',
    'USAJobs.gov':               'https://www.usajobs.gov/',
}

# Job board → SOC mapping（手动定义）
JB_SOC_MAP = {
    'MaineDOT Careers':         ['17-2051.01','47-2073','47-2061','47-1011'],
    'Maine Turnpike Authority':  ['47-2073','17-3022','47-1011'],
    'Maine.gov Job Board':       ['17-2051','13-1082','19-4044'],
    'AGC Maine':                 ['47-2031','47-2061','47-2111','47-1011','11-9021'],
    'NSPE-Maine':                ['17-2051','17-2051.01','17-3022'],
    'Cianbro':                   ['47-2221','47-2111','47-2073','47-2061'],
    'Sargent':                   ['47-2073','47-1011','47-2071','47-2061'],
    'Reed & Reed':               ['17-2051','53-7021','47-2221'],
    'Sebago Technics':           ['17-2051','17-3022'],
    'Haley Ward':                ['17-2051','19-4044'],
    'Live-and-Work-in-Maine':    ['11-9021','13-1082','11-9041'],
    'Maine Hire-A-Vet':          ['47-2061','47-2031','47-1011'],
    'ConstructionJobs.com':      list(ALL_SOC.keys()),
    'iHireConstruction':         ['11-9021','13-1082','47-1011','47-2031','47-2061'],
    'ASCE Career Connections':   ['17-2051','17-2051.01','17-3022'],
    'EngineerJobs.com':          ['17-2051','17-2051.01','11-9041'],
    'CMAA Career HQ':            ['11-9021','13-1082'],
    'AASHTO Jobs':               ['17-2051.01','17-2051','17-3022'],
    'ITE Hire':                  ['17-2051.01','17-2051'],
    'RoadTechs':                 ['47-2073','47-2061','47-2071','47-2221'],
    'Michael Page Construction': ['11-9021','11-9041','13-1082'],
    'USAJobs.gov':               ['17-2051.01','17-2051','11-9041'],
}

for board_name, soc_codes in JB_SOC_MAP.items():
    url = jb_urls.get(board_name)
    scope = 'maine' if board_name in [
        'MaineDOT Careers','Maine Turnpike Authority','Maine.gov Job Board',
        'AGC Maine','NSPE-Maine','Cianbro','Sargent','Reed & Reed',
        'Sebago Technics','Haley Ward','Live-and-Work-in-Maine','Maine Hire-A-Vet'
    ] else 'national'

    for soc in soc_codes:
        occ_title = ALL_SOC.get(soc, '')
        if not url:
            url = f"https://www.indeed.com/jobs?q={urllib.parse.quote(occ_title)}&l=Maine"
        records.append({
            'soc_code':      soc,
            'occ_title':     occ_title,
            'employer_name': board_name,
            'agc_type':      scope,
            'url':           url,
            'url_type':      'job_board',
            'source_type':   'job_board',
        })

# ── 6. 输出 ───────────────────────────────────────────────────────────────────
df_out = pd.DataFrame(records)
df_out = df_out.drop_duplicates(subset=['soc_code','employer_name'])
df_out = df_out.sort_values(['soc_code','source_type','employer_name'])

print(f"\nFinal: {len(df_out)} rows, {df_out['soc_code'].nunique()} SOC codes covered")
print(df_out.groupby('source_type')['employer_name'].count())
print("\nSample:")
print(df_out.head(10).to_string())

df_out.to_csv('Output_data/employers.csv', index=False)
print("\n✅ employers.csv saved.")