import pandas as pd
from openpyxl import load_workbook

EXCEL_PATH  = 'Download_data/Career Drive Project Data Sources.xlsx'
MASTER_PATH = 'Output_data/occupations_master.csv'
OUTPUT_PATH = 'Output_data/training_programs.csv'

# ── 1/5: Extract Community College Programs ───────────────────────────────────
wb = load_workbook(EXCEL_PATH)
ws = wb['Community College Programs']

# Extract rows 3-16 (Community College Programs only)
cc_rows = []
for row in ws.iter_rows(min_row=3, max_row=16, min_col=1, max_col=4):
    college    = row[0].value
    program    = row[1].value
    credential = row[2].value
    url        = row[3].hyperlink.target if row[3].hyperlink else None
    cc_rows.append({
        'college':      college,
        'program_name': program,
        'credential':   credential,
        'url':          url,
    })

cc = pd.DataFrame(cc_rows)
cc = cc.dropna(subset=['program_name'])

# Normalize credential to program_type
def map_program_type(credential):
    if credential is None:
        return 'unknown'
    c = credential.lower()
    if 'short' in c:
        return 'certificate'
    elif 'aas' in c and 'certificate' in c:
        return 'associate_or_certificate'
    elif 'aas' in c:
        return 'associate'
    elif 'certificate' in c:
        return 'certificate'
    return 'unknown'

cc['program_type'] = cc['credential'].apply(map_program_type)
cc['source'] = 'community_college'

# Expand college abbreviations to full names
college_names = {
    'CMCC': 'Central Maine Community College',
    'EMCC': 'Eastern Maine Community College',
    'KVCC': 'Kennebec Valley Community College',
    'NMCC': 'Northern Maine Community College',
    'SMCC': 'Southern Maine Community College',
    'WCCC': 'Washington County Community College',
    'YCCC': 'York County Community College',
}
cc['college'] = cc['college'].map(college_names).fillna(cc['college'])

# ── 2/5: Extract Workforce Initiatives ───────────────────────────────────────
init_rows = []
for row in ws.iter_rows(min_row=21, max_row=26, min_col=1, max_col=4):
    name    = row[0].value
    focus   = row[1].value
    best_for = row[2].value
    url     = row[3].hyperlink.target if row[3].hyperlink else None
    init_rows.append({
        'program_name': name,
        'focus':        focus,
        'best_for':     best_for,
        'url':          url,
    })

initiatives = pd.DataFrame(init_rows)
initiatives = initiatives.dropna(subset=['program_name'])

# ── 2.5/5: Extract UMaine Programs ───────────────────────────────────────────
ws_umaine = wb['Umaine Programs']

umaine_rows = []
for row in ws_umaine.iter_rows(min_row=4, max_row=13, min_col=1, max_col=5):
    campus   = row[0].value
    category = row[1].value
    degree   = row[2].value
    program  = row[3].value
    url      = row[4].hyperlink.target if row[4].hyperlink else None
    umaine_rows.append({
        'college':      campus,
        'program_name': program,
        'credential':   degree,
        'program_type': 'university',
        'url':          url,
        'source':       'umaine',
    })

umaine = pd.DataFrame(umaine_rows)
umaine = umaine.dropna(subset=['program_name'])

# Clean Google-wrapped URLs
umaine['url'] = umaine['url'].str.replace(
    r'https://www\.google\.com/search\?q=', '', regex=True
)

print(f"\n2.5/5 ── UMaine Programs: {len(umaine)} rows extracted")
print(umaine[['college', 'program_name', 'credential', 'url']].to_string())

# ── 3/5: Extract AGC Apprenticeships ─────────────────────────────────────────
app_raw = pd.read_excel(EXCEL_PATH, sheet_name='AGC Sponsored Apprenticeships ',
                        header=1, dtype=str)
app_raw.columns = ['rapids_code', 'program_name', 'term_hours']
app_raw = app_raw.dropna(subset=['program_name'])

# Remove non-construction trades
exclude = ['Arborist', 'Lead Logging Equipment Operator']
app = app_raw[~app_raw['program_name'].isin(exclude)].copy()

app['college']       = 'AGC Maine'
app['credential']    = 'Apprenticeship'
app['program_type']  = 'apprenticeship'
app['url']           = 'https://buildingmaine.com/pre-apprenticeship'
app['source']        = 'agc_apprenticeship'

# ── 4/5: SOC Mapping ──────────────────────────────────────────────────────────
# Load occupation master to get occ_title for each soc_code
occ = pd.read_csv(MASTER_PATH)[['soc_code', 'occ_title']]

# Manual mapping: program_name → list of soc_codes
soc_mapping = {
    # Community College Programs
    'Building Construction Technology':           ['47-2031'],
    'HVAC/R Technology':                          ['47-2152'],
    'Plumbing & Heating Technology':              ['47-2152'],
    'Building Construction':                      ['47-2031'],
    'Plumbing':                                   ['47-2152'],
    'Refrigeration, AC & Heating':                ['47-2152'],
    'Construction Technology':                    ['47-2031'],
    'Heavy Equipment Operations':                 ['47-2073'],
    'Heavy Equipment Maintenance':                ['47-2073'],
    'Industrial Welding Techniques':              ['47-2221'],
    # Workforce Initiatives
    'Maine Construction Academy (MCA)':           ['47-2031', '47-2071', '47-2081', '47-2171', '47-4021', '47-4061'],
    'Alfond Center: Trades Track':                ['47-2152'],
    'TREC Program':                               ['47-2152'],
    'Transfer ME (Pre-Engineering)':              ['17-2051'],
    'Maine Workforce Development Compact':        ['11-9021'],
    'Advanced Technology Centers':                ['17-2051'],
    # UMaine Programs
    'Construction Engineering Technology': ['11-9021', '17-2051'],
    'Architecture (5-Year Professional Degree)': [],
    'Civil & Environmental Engineering':   ['17-2051', '17-2051.01'],
    'Mechanical Engineering':              ['11-9041'],
    'Electrical & Computer Engineering':   ['47-2111'],
    'Surveying Engineering Technology':    ['17-2051'],
    'Industrial Engineering':              ['11-9041'],
    'Maine Engineering Pathways':          ['17-2051', '17-2051.01'],
    'Bridge Carpenter/Heavy Highway':             ['47-2031'],
    'Construction Carpenter':                     ['47-2031'],
    'Construction Craft Concrete Laborer':        ['47-2031'],
    'Construction Craft Heavy / Highway Laborer': ['47-4051'],
    'Construction Equipment Operator':            ['47-2073'],
    'Construction Specialist':                    ['47-2031'],
    'Crane Mechanic':                             ['47-2073'],
    'Crane Operator':                             ['47-2072'],
    'Earthworks Laborer':                         ['47-2151'],
    'Electrician':                                ['47-2111'],
    'Fencing Installer':                          ['47-2031'],
    'Firestopping Installer':                     ['47-2132'],
    'Firestopping Technician':                    ['47-2132'],
    'Foreman':                                    ['47-1011'],
    'Marine Carpenter - Heavy Civil':             ['47-2031'],
    'Solar Mechanical Installation Technician':   ['47-2152'],
    'Welder':                                     ['47-2221'],
}

# Expand mapping into rows
mapping_rows = [
    {'program_name': prog, 'soc_code': soc}
    for prog, codes in soc_mapping.items()
    for soc in codes
]
mapping_df = pd.DataFrame(mapping_rows)

# ── 5/5: Merge all sources and output ────────────────────────────────────────
# Standardize columns across all three sources
initiatives['college']      = 'Workforce Initiative'
initiatives['credential']   = initiatives['focus']
initiatives['program_type'] = 'bootcamp'
initiatives['source']       = 'workforce_initiative'
initiatives = initiatives.drop(columns=['focus', 'best_for'])

# Combine all three sources
all_programs = pd.concat([cc, initiatives, umaine, app], ignore_index=True)

# Join SOC mapping
all_programs = all_programs.merge(mapping_df, on='program_name', how='inner')

# Join occ_title from occupation master
all_programs = all_programs.merge(occ, on='soc_code', how='left')

# Final column order
all_programs = all_programs[[
    'soc_code', 'occ_title', 'program_name', 'college',
    'program_type', 'credential', 'url', 'source'
]]

all_programs = all_programs.sort_values(['soc_code', 'program_type'])
all_programs.to_csv(OUTPUT_PATH, index=False)

print(f"\n5/5 ── training_programs.csv saved: {len(all_programs)} rows")
print(f"       Unique occupations covered: {all_programs['soc_code'].nunique()}")
print(f"       Unique programs: {all_programs['program_name'].nunique()}")
print(f"\n── Coverage Check ──")
all_soc = set(occ['soc_code'])
covered = set(all_programs['soc_code'])
missing = all_soc - covered
if missing:
    print(f"Occupations with NO training program: {missing}")
else:
    print("All occupations have at least one training program ✓")
print(f"\n── Preview ──")
print(all_programs[['soc_code','occ_title','program_name','program_type']].head(10).to_string())